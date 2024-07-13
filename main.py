import json
import os
import re
from datetime import datetime
import openpyxl
from fpdf import FPDF


def read_log_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        return file.read()


def parse_laravel_log(log_contents):
    log_pattern = re.compile(r'\[(.*?)\] (\w+)\.(\w+): (.*?)(?=\[\d{4}-\d{2}-\d{2}|\Z)', re.DOTALL)
    log_entries = log_pattern.findall(log_contents)

    parsed_logs = []
    for entry in log_entries:
        parsed_log = parse_log_entry(entry)
        parsed_logs.append(parsed_log)

    return sorted(parsed_logs, key=lambda x: x["timestamp"])


def parse_log_entry(entry):
    timestamp, environment, log_level, details = entry
    details_pattern = re.compile(r'(.*?)\n\[stacktrace\]\n(.*)', re.DOTALL)
    match = details_pattern.match(details.strip())

    if match:
        message, stack_trace = match.groups()
    else:
        message, stack_trace = details.strip(), ''

    return {
        "timestamp": datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S'),
        "environment": environment.lower(),
        "log_level": log_level,
        "message": message.strip(),
        "stack_trace": stack_trace.strip()
    }


def generate_json_report(parsed_logs, output_file):
    report = []
    for log in parsed_logs:
        report_entry = {
            "Hata Zamanı": log["timestamp"].strftime('%Y-%m-%d %H:%M:%S'),
            "Çevre": log["environment"],
            "Hata Seviyesi": log["log_level"],
            "Hata Mesajı": log["message"],
            "Yığın İzleme": log["stack_trace"]
        }
        report.append(report_entry)

    with open(output_file, 'w', encoding='utf-8') as file:
        json.dump(report, file, ensure_ascii=False, indent=4)


def generate_text_report(parsed_logs, output_file):
    with open(output_file, 'w', encoding='utf-8') as file:
        for log in parsed_logs:
            file.write(f"Hata Zamanı: {log['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}\n")
            file.write(f"Çevre: {log['environment']}\n")
            file.write(f"Hata Seviyesi: {log['log_level']}\n")
            file.write(f"Hata Mesajı: {log['message']}\n")
            file.write(f"Yığın İzleme: {log['stack_trace']}\n")
            file.write("-" * 50 + "\n")


def generate_excel_report(parsed_logs, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Hata Zamanı", "Çevre", "Hata Seviyesi", "Hata Mesajı", "Yığın İzleme"])

    for log in parsed_logs:
        ws.append([
            log['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
            log['environment'],
            log['log_level'],
            log['message'],
            log['stack_trace']
        ])

    wb.save(output_file)


def generate_pdf_report(parsed_logs, output_file):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    for log in parsed_logs:
        pdf.cell(200, 10, txt=f"Hata Zamanı: {log['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
        pdf.cell(200, 10, txt=f"Çevre: {log['environment']}", ln=True)
        pdf.cell(200, 10, txt=f"Hata Seviyesi: {log['log_level']}", ln=True)
        pdf.multi_cell(0, 10, txt=f"Hata Mesajı: {log['message']}")
        pdf.multi_cell(0, 10, txt=f"Yığın İzleme: {log['stack_trace']}")
        pdf.cell(200, 10, txt="-" * 50, ln=True)

    pdf.output(output_file)


def deletedLogCheckControl(file_path):
    deletedControl = int(input('Log dosyası silinsin istiyor musunuz?: ( 1 sil,2 silme varsayılan ) ').strip() or "2")

    if deletedControl == 1:
        os.remove(file_path)
        print(f"Deleted processed log file: {file_path}")


def analyze_error_logs(log_directory, output_directory, output_format):
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    for filename in os.listdir(log_directory):
        if filename.endswith('.log'):
            file_path = os.path.join(log_directory, filename)
            try:
                log_contents = read_log_file(file_path)
                parsed_logs = parse_laravel_log(log_contents)

                report_filename = f"report_{os.path.splitext(filename)[0]}.{output_format}"
                report_path = os.path.join(output_directory, report_filename)

                if output_format == 'json':
                    generate_json_report(parsed_logs, report_path)
                elif output_format == 'txt':
                    generate_text_report(parsed_logs, report_path)
                elif output_format == 'xlsx':
                    generate_excel_report(parsed_logs, report_path)
                elif output_format == 'pdf':
                    generate_pdf_report(parsed_logs, report_path)

                print(f"Report generated for {filename}: {report_path}")

                deletedLogCheckControl(file_path)

            except Exception as e:
                print(f"Error processing {filename}: {str(e)}")


if __name__ == "__main__":
    log_directory = "errors"
    output_directory = "output"

    print("Lütfen çıktı formatını seçin:")
    print("1. JSON")
    print("2. Text")
    print("3. Excel")
    print("4. PDF ( Yakında )")

    choice = input("Seçiminiz (1/2/3/4): ")

    format_map = {'1': 'json', '2': 'txt', '3': 'xlsx', '4': 'pdf'}
    output_format = format_map.get(choice, 'json')

    analyze_error_logs(log_directory, output_directory, output_format)
